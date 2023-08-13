import openpyxl

# Tạo một workbook mới
workbook = openpyxl.Workbook()

# Chọn sheet đầu tiên
sheet = workbook.active

# Đặt tên cho các cột
sheet['A1'] = 'Họ và tên'
sheet['B1'] = 'Lương'
sheet['C1'] = 'Thưởng'

# Dữ liệu lương nhân viên
employees = [
    {'name': 'Nguyễn Văn A', 'salary': 5000, 'bonus': 1000},
    {'name': 'Trần Thị B', 'salary': 6000, 'bonus': 1500},
    {'name': 'Lê Văn C', 'salary': 4500, 'bonus': 800},
    # Thêm nhân viên khác tại đây...
]

# Ghi dữ liệu lương vào bảng
row = 2  # Dòng bắt đầu ghi dữ liệu
for employee in employees:
    sheet.cell(row=row, column=1, value=employee['name'])
    sheet.cell(row=row, column=2, value=employee['salary'])
    sheet.cell(row=row, column=3, value=employee['bonus'])
    row += 1

# Lưu workbook
workbook.save('luong_nhan_vien.xlsx')
